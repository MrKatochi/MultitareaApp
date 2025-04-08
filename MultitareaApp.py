import os
import shutil
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import openpyxl
import logging
from threading import Thread
from ttkthemes import ThemedStyle
import datetime
import mysql.connector
import paramiko
from openpyxl.utils import get_column_letter


class MultitareaApp:
    TEMAS = {
        "Por Defecto": "default", "Adapta": "adapta", "Aquativo": "aquativo",
        "Arc": "arc", "Black": "black", "Blue": "blue", "Breeze": "breeze",
        "Clearlooks": "clearlooks", "Elegance": "elegance", "Equilux": "equilux",
        "ITFT1": "itft1", "Keramik": "keramik", "Kroc": "kroc",
        "Plastik": "plastik", "Radiance (Ubuntu)": "radiance",
        "Scid themes": "scid", "Smog": "smog",
        "winxpblue": "winxpblue", "yaru": "yaru"
    }

    CATEGORIAS = {
        'Imágenes': ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff', '.webp'],
        'Documentos': ['.doc', '.docx', '.pdf', '.txt', '.rtf', '.odt', '.xls', '.xlsx', '.ppt', '.pptx'],
        'Audio': ['.mp3', '.wav', '.flac', '.m4a', '.aac', '.ogg'],
        'Video': ['.mp4', '.avi', '.mov', '.wmv', '.mkv', '.flv', '.webm'],
        'Archivos Comprimidos': ['.zip', '.rar', '.7z', '.tar', '.gz'],
        'Ejecutables': ['.exe', '.msi'], 'Bases de Datos': ['.sql']
    }

    def __init__(self, master):
        self.master = master
        self.master.title("Herramientas")
        self.master.geometry("800x600")
        self.master.resizable(False, False)

        # Configurar logging en la carpeta del proyecto
        log_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'app.log')
        logging.basicConfig(filename=log_file, level=logging.INFO,
                            format='%(asctime)s - %(levelname)s - %(message)s')

        # Inicializar estilo y variables
        self.estilo = ThemedStyle(self.master)
        self.estilo.set_theme("default")
        self.directorio_seleccionado = None
        self.historial_operaciones = []
        self.botones_organizador = []

        # Crear pestañas e interfaz
        self.tab_control = ttk.Notebook(self.master)
        for nombre, metodo in [("Organizador", self.crear_tab_organizador),
                               ("Base de Datos", self.crear_tab_base_datos),
                               ("Configuración", self.crear_tab_conf)]:
            frame = ttk.Frame(self.tab_control)
            self.tab_control.add(frame, text=nombre)
            metodo(frame)

        self.tab_control.pack(expand=1, fill="both")
        ttk.Label(self.master, text="© 2025 Todos los derechos reservados",
                  anchor=tk.CENTER).pack(side=tk.BOTTOM, fill=tk.X)

    def crear_tab_organizador(self, frame):
        ttk.Button(frame, text="Seleccionar Directorio", command=self.seleccionar_directorio).grid(row=0, column=0, padx=5, pady=5)
        for i, (texto, comando) in enumerate([
            ("Organizar Archivos", self.organizar_archivos),
            ("Deshacer", self.deshacer_operacion),
            ("Generar Excel", self.crear_excel_archivos)], start=1):

            boton = ttk.Button(frame, text=texto, command=comando, state=tk.DISABLED)
            boton.grid(row=0, column=i, padx=5, pady=5)
            self.botones_organizador.append(boton)

        # Barra de progreso y estado
        self.barra_progreso = ttk.Progressbar(frame, orient="horizontal", length=400, mode="determinate")
        self.barra_progreso.grid(row=1, column=0, columnspan=4, pady=10, sticky='ew')

        self.etiqueta_estado = ttk.Label(frame, text="Estado: Esperando operación")
        self.etiqueta_estado.grid(row=2, column=0, columnspan=4)

    def crear_tab_base_datos(self, frame):
        db_frame = tk.LabelFrame(frame, text="Conexión a Base de Datos", padx=10, pady=10)
        db_frame.grid(row=0, column=0, padx=10, pady=10, sticky=tk.N)

        # Variables para almacenar información de conexión
        self.host_var = tk.StringVar(value="localhost")
        self.ssh_tunnel_var = tk.BooleanVar(value=False)

        # Checkbox para conexión SSH
        check_ssh = tk.Checkbutton(db_frame, text="Conexión SSH", variable=self.ssh_tunnel_var, command=self.mostrar_ocultar_ssh)
        check_ssh.grid(row=0, column=0, sticky=tk.W)

        # Campos para conexión
        tk.Label(db_frame, text="Host:").grid(row=1, column=0, sticky=tk.W)
        self.entry_host = tk.Entry(db_frame, width=30, textvariable=self.host_var)
        self.entry_host.grid(row=1, column=1, padx=5, pady=2)

        tk.Label(db_frame, text="Puerto:").grid(row=1, column=2, sticky=tk.W)
        self.entry_port = tk.Entry(db_frame, width=10)
        self.entry_port.insert(0, "3306")
        self.entry_port.grid(row=1, column=3, padx=5, pady=2)

        tk.Label(db_frame, text="Usuario:").grid(row=2, column=0, sticky=tk.W)
        self.entry_user = tk.Entry(db_frame, width=30)
        self.entry_user.grid(row=2, column=1, padx=5, pady=2)

        tk.Label(db_frame, text="Contraseña:").grid(row=2, column=2, sticky=tk.W)
        self.entry_password = tk.Entry(db_frame, width=30, show="*")
        self.entry_password.grid(row=2, column=3, padx=5, pady=2)

        tk.Label(db_frame, text="Base de Datos:").grid(row=3, column=0, sticky=tk.W)
        self.entry_dbname = tk.Entry(db_frame, width=30)
        self.entry_dbname.grid(row=3, column=1, padx=5, pady=2)

        # Frame SSH (oculto inicialmente)
        self.ssh_frame = tk.LabelFrame(db_frame, text="Configuración SSH", padx=5, pady=5)
        self.ssh_frame.grid(row=4, column=0, columnspan=4, sticky=tk.W, padx=5, pady=5)
        self.ssh_frame.grid_remove()  # Ocultar inicialmente

        tk.Label(self.ssh_frame, text="SSH Host:").grid(row=0, column=0, sticky=tk.W)
        self.entry_ssh_host = tk.Entry(self.ssh_frame, width=30)
        self.entry_ssh_host.grid(row=0, column=1, padx=5, pady=2)

        tk.Label(self.ssh_frame, text="SSH Usuario:").grid(row=1, column=0, sticky=tk.W)
        self.entry_ssh_user = tk.Entry(self.ssh_frame, width=30)
        self.entry_ssh_user.grid(row=1, column=1, padx=5, pady=2)

        tk.Label(self.ssh_frame, text="SSH Contraseña:").grid(row=2, column=0, sticky=tk.W)
        self.entry_ssh_password = tk.Entry(self.ssh_frame, width=30, show="*")
        self.entry_ssh_password.grid(row=2, column=1, padx=5, pady=2)

        # Botón para conectar
        self.boton_conectar = tk.Button(db_frame, text="Conectar", command=self.conectar_base_datos)
        self.boton_conectar.grid(row=7, columnspan=4, pady=10)

        # Desplegable para tablas y tabla para registros
        tk.Label(frame, text="Tablas:").grid(row=1, column=0, padx=10, pady=5, sticky=tk.W)
        self.combo_tablas = ttk.Combobox(frame, state="readonly")
        self.combo_tablas.grid(row=1, column=1, padx=10, pady=5, sticky=tk.W)
        self.combo_tablas.bind("<<ComboboxSelected>>", self.mostrar_registros_tabla)

        # Treeview para mostrar registros
        '''
        self.tree_registros = ttk.Treeview(frame, columns=("ID", "Nombre", "Valor"), show="headings")
        self.tree_registros.heading("ID", text="ID")
        self.tree_registros.heading("Nombre", text="Nombre")
        self.tree_registros.heading("Valor", text="Valor")
        self.tree_registros.grid(row=2, column=0, columnspan=2, padx=10, pady=5, sticky=tk.NSEW)
        frame.columnconfigure(0, weight=1)
        frame.rowconfigure(2, weight=1)
        '''
        # Treeview con scroll horizontal y vertical
        tree_frame = ttk.Frame(frame)
        tree_frame.grid(row=2, column=0, columnspan=2, padx=10, pady=5, sticky=tk.NSEW)
        frame.columnconfigure(0, weight=1)
        frame.rowconfigure(2, weight=1)

        scroll_y = ttk.Scrollbar(tree_frame, orient="vertical")
        scroll_x = ttk.Scrollbar(tree_frame, orient="horizontal")

        self.tree_registros = ttk.Treeview(
            tree_frame,
            yscrollcommand=scroll_y.set,
            xscrollcommand=scroll_x.set,
            show="headings"
        )

        scroll_y.config(command=self.tree_registros.yview)
        scroll_x.config(command=self.tree_registros.xview)

        self.tree_registros.grid(row=0, column=0, sticky="nsew")
        scroll_y.grid(row=0, column=1, sticky="ns")
        scroll_x.grid(row=1, column=0, sticky="ew")

        tree_frame.columnconfigure(0, weight=1)
        tree_frame.rowconfigure(0, weight=1)

    def crear_tab_conf(self, frame):
        conf_frame = ttk.Frame(frame, padding=10)
        conf_frame.grid(row=0, column=0, sticky='nsew')

        # Selector de temas (Combobox)
        self.var_tema = tk.StringVar(value="default")
        self.combo_tema = ttk.Combobox(conf_frame, textvariable=self.var_tema, values=list(self.TEMAS.keys()),
                                       state="readonly")
        self.combo_tema.grid(row=0, column=0, sticky='w', pady=5)
        self.combo_tema.bind("<<ComboboxSelected>>", self.cambiar_tema)
        self.combo_tema.set("Por Defecto")

    def cambiar_tema(self, event=None):
        tema_nombre = self.combo_tema.get()
        tema = self.TEMAS[tema_nombre]
        try:
            self.estilo.set_theme(tema)
            logging.info(f"Tema cambiado a: {tema}")
        except Exception as e:
            logging.error(f"Error al cambiar tema: {tema} - {e}")
            messagebox.showerror("Error", f"No se pudo aplicar el tema: {tema}")
            self.estilo.set_theme("default")

    def seleccionar_directorio(self):
        self.directorio_seleccionado = filedialog.askdirectory()
        if self.directorio_seleccionado:
            self.habilitar_botones()
            self.etiqueta_estado.config(text=f"Directorio seleccionado: {self.directorio_seleccionado}")
            logging.info(f"Directorio seleccionado: {self.directorio_seleccionado}")

    def habilitar_botones(self):
        frame = self.tab_control.winfo_children()[0].winfo_children()[0]
        for btn in frame.winfo_children()[1:]:  # Skip "Seleccionar Directorio" button
            btn.config(state=tk.NORMAL)

    def mostrar_ocultar_ssh(self):
        if self.ssh_tunnel_var.get():
            self.ssh_frame.grid()
        else:
            self.ssh_frame.grid_remove()

    def organizar_archivos(self):
        if not self.directorio_seleccionado:
            messagebox.showwarning("Error", "Seleccione un directorio primero")
            return

        self.etiqueta_estado.config(text="Organizando archivos...")
        self.barra_progreso['value'] = 0  # Reiniciar progreso
        self.master.update_idletasks()    # Refrescar GUI
        Thread(target=self._organizar_archivos).start()


    def _organizar_archivos(self):
        try:
            self.historial_operaciones.append([])
            total_archivos = sum(
                1 for archivo in os.listdir(self.directorio_seleccionado)
                if os.path.isfile(os.path.join(self.directorio_seleccionado, archivo))
            )

            progreso = 0

            for categoria, extensiones in self.CATEGORIAS.items():
                categoria_dir = os.path.join(self.directorio_seleccionado, categoria)
                os.makedirs(categoria_dir, exist_ok=True)

                for archivo in os.listdir(self.directorio_seleccionado):
                    archivo_path = os.path.join(self.directorio_seleccionado, archivo)
                    if os.path.isfile(archivo_path):
                        if any(archivo.lower().endswith(ext) for ext in extensiones):
                            destino = os.path.join(categoria_dir, archivo)
                            if archivo_path != destino:
                                shutil.move(archivo_path, destino)
                                self.historial_operaciones[-1].append((destino, archivo_path))
                                progreso += 1
                                self.actualizar_progreso((progreso / total_archivos) * 100)
            
            self.actualizar_progreso(100)  # Asegura barra llena al final
            self.finalizar_operacion("Organización completada con éxito")
        except Exception as e:
            logging.error(f"Error al organizar archivos: {str(e)}")
            messagebox.showerror("Error", f"Error al organizar archivos: {str(e)}")
        

    def deshacer_operacion(self):
        if not self.historial_operaciones:
            messagebox.showwarning("Advertencia", "No hay operaciones para deshacer")
            return

        try:
            ultima_operacion = self.historial_operaciones.pop()
            for destino, origen in reversed(ultima_operacion):
                if os.path.exists(destino):
                    shutil.move(destino, origen)

            # Eliminar carpetas vacías
            for categoria in self.CATEGORIAS:
                carpeta = os.path.join(self.directorio_seleccionado, categoria)
                if os.path.exists(carpeta) and not os.listdir(carpeta):
                    os.rmdir(carpeta)

            self.finalizar_operacion("Operación deshecha correctamente")
        except Exception as e:
            logging.error(f"Error al deshacer: {str(e)}")
            messagebox.showerror("Error", f"Error al deshacer: {str(e)}")

    def crear_excel_archivos(self):
        if not self.directorio_seleccionado:
            messagebox.showwarning("Error", "Seleccione un directorio primero")
            return

        try:
            workbook = openpyxl.Workbook()
            workbook.remove(workbook.active)

            for categoria in self.CATEGORIAS:
                categoria_dir = os.path.join(self.directorio_seleccionado, categoria)
                if not os.path.exists(categoria_dir):
                    continue

                hoja = workbook.create_sheet(title=categoria[:31])
                hoja.append(["Nombre", "Ruta", "Tipo", "Tamaño (MB)", "Creado", "Modificado"])

                for root, dirs, files in os.walk(categoria_dir):
                    for d in dirs:
                        self._agregar_registro(hoja, os.path.join(root, d), True)
                    for f in files:
                        self._agregar_registro(hoja, os.path.join(root, f), False)

                self._autoajustar_columnas(hoja)

            excel_path = os.path.join(self.directorio_seleccionado,
                                    f"Reporte_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")
            workbook.save(excel_path)
            messagebox.showinfo("Éxito", f"Reporte generado:\n{excel_path}")
        except Exception as e:
            logging.error(f"Error al generar Excel: {str(e)}")
            messagebox.showerror("Error", f"Error al generar Excel: {str(e)}")

    def _agregar_registro(self, hoja, ruta, es_carpeta):
        nombre = os.path.basename(ruta)
        tamaño = "-"
        creado = datetime.datetime.fromtimestamp(os.path.getctime(ruta)).strftime('%Y-%m-%d %H:%M')
        modificado = datetime.datetime.fromtimestamp(os.path.getmtime(ruta)).strftime('%Y-%m-%d %H:%M')

        if not es_carpeta:
            tamaño = round(os.path.getsize(ruta) / (1024 * 1024), 2)

        tipo = "Carpeta" if es_carpeta else os.path.splitext(ruta)[1].upper()
        hoja.append([nombre, ruta, tipo, tamaño, creado, modificado])

    def _autoajustar_columnas(self, hoja):
        for columna in hoja.columns:
            max_longitud = 0
            columna_letra = get_column_letter(columna[0].column)
            for celda in columna:
                try:
                    if len(str(celda.value)) > max_longitud:
                        max_longitud = len(str(celda.value))
                except:
                    pass
            longitud_ajustada = (max_longitud + 2)
            hoja.column_dimensions[columna_letra].width = longitud_ajustada

    def conectar_base_datos(self):
        host = self.entry_host.get()
        port = int(self.entry_port.get())
        user = self.entry_user.get()
        password = self.entry_password.get()
        db_name = self.entry_dbname.get()
        ssh_tunnel = self.ssh_tunnel_var.get()

        try:
            if ssh_tunnel:
                ssh_host = self.entry_ssh_host.get()
                ssh_user = self.entry_ssh_user.get()
                ssh_password = self.entry_ssh_password.get()

                # Configurar el túnel SSH
                self.ssh_client = paramiko.SSHClient()
                self.ssh_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
                self.ssh_client.connect(ssh_host, username=ssh_user, password=ssh_password)

                transport = self.ssh_client.get_transport()
                transport.request_port_forward('', port)

                self.db_connection = mysql.connector.connect(
                    host='localhost',  # Conexión a través del túnel SSH
                    port=port,
                    user=user,
                    password=password,
                    database=db_name
                )
            else:
                self.db_connection = mysql.connector.connect(
                    host=host,
                    port=port,
                    user=user,
                    password=password,
                    database=db_name
                )

            self.cursor = self.db_connection.cursor()
            messagebox.showinfo("Conexión Exitosa", "Conexión a la base de datos establecida.")
            self.cargar_tablas()

        except mysql.connector.Error as err:
            logging.error(f"Error al conectar a la base de datos: {err}")
            messagebox.showerror("Error de Conexión", f"Error: {err}")
        except paramiko.AuthenticationException:
            logging.error("Error de autenticación SSH.")
            messagebox.showerror("Error", "Error de autenticación SSH.")
        except Exception as e:
            logging.error(f"Error inesperado: {e}")
            messagebox.showerror("Error", f"Error inesperado: {e}")

    def cargar_tablas(self):
        if self.db_connection:
            try:
                self.cursor.execute("SHOW TABLES")
                tablas = [tabla[0] for tabla in self.cursor.fetchall()]
                self.combo_tablas['values'] = tablas
            except mysql.connector.Error as err:
                logging.error(f"Error al cargar las tablas: {err}")
                messagebox.showerror("Error", f"Error al cargar las tablas: {err}")

    def habilitar_botones(self):
        for boton in self.botones_organizador:
            boton.config(state=tk.NORMAL)

    def mostrar_registros_tabla(self, event=None):
        tabla_seleccionada = self.combo_tablas.get()
        if tabla_seleccionada:
            try:
                self.cursor.execute(f"SELECT * FROM {tabla_seleccionada}")
                registros = self.cursor.fetchall()
                columnas = [desc[0] for desc in self.cursor.description]

                # Limpiar columnas y registros actuales
                self.tree_registros.delete(*self.tree_registros.get_children())
                self.tree_registros["columns"] = columnas

                for col in columnas:
                    self.tree_registros.heading(col, text=col)
                    self.tree_registros.column(col, width=100, anchor=tk.W)  # Puedes ajustar el ancho si lo deseas

                # Insertar registros
                for registro in registros:
                    self.tree_registros.insert("", tk.END, values=registro)

            except mysql.connector.Error as err:
                logging.error(f"Error al mostrar registros de la tabla: {err}")
                messagebox.showerror("Error", f"Error al mostrar registros de la tabla: {err}")


    def actualizar_progreso(self, progreso):
        self.barra_progreso['value'] = progreso
        self.master.update_idletasks()

    def finalizar_operacion(self, mensaje):
        self.etiqueta_estado.config(text=f"Estado: {mensaje}")
        messagebox.showinfo("Información", mensaje)
        logging.info(mensaje)

if __name__ == "__main__":
    root = tk.Tk()
    app = MultitareaApp(root)
    root.mainloop()
